"""
Excel export utility for restaurant data
Creates well-formatted Excel files with multiple sheets
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


def create_restaurants_excel(restaurants, state="Lagos"):
    """
    Create a well-formatted Excel workbook with restaurant data

    Args:
        restaurants: List of restaurant dictionaries
        state: State name for filtering (default: Lagos)

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Create Summary Sheet
    create_summary_sheet(wb, restaurants, state)

    # 2. Create All Restaurants Sheet
    create_all_restaurants_sheet(wb, restaurants, state)

    # 3. Create sheets by LGA
    create_lga_sheets(wb, restaurants, state)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def create_summary_sheet(wb, restaurants, state):
    """Create summary statistics sheet"""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = f"{state} Restaurants Database"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30

    # Metadata
    row = 3
    ws[f'A{row}'] = "Generated:"
    ws[f'B{row}'] = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "Total Restaurants:"
    ws[f'B{row}'] = len(restaurants)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(size=12, bold=True, color="1F4788")

    # Statistics by LGA
    row += 2
    ws[f'A{row}'] = "Distribution by LGA"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    # Headers
    ws[f'A{row}'] = "LGA"
    ws[f'B{row}'] = "Restaurants"
    ws[f'C{row}'] = "Percentage"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Count by LGA
    lga_counts = {}
    for r in restaurants:
        lga = r.get('lga') or 'Unknown'
        lga_counts[lga] = lga_counts.get(lga, 0) + 1

    row += 1
    start_data_row = row
    # Sort with None-safe comparison
    sorted_lgas = sorted(lga_counts.keys(), key=lambda x: x if x else 'ZZZ')
    for lga in sorted_lgas:
        count = lga_counts[lga]
        percentage = (count / len(restaurants)) * 100

        ws[f'A{row}'] = lga
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{percentage:.1f}%"
        ws[f'B{row}'].alignment = Alignment(horizontal="center")
        ws[f'C{row}'].alignment = Alignment(horizontal="center")
        row += 1

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r in range(start_data_row - 1, row):
        for col in ['A', 'B', 'C']:
            ws[f'{col}{r}'].border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_all_restaurants_sheet(wb, restaurants, state):
    """Create sheet with all restaurants"""
    ws = wb.create_sheet(f"All {state} Restaurants")

    # Headers
    headers = ['#', 'Restaurant Name', 'City', 'LGA', 'Address', 'Rating', 'Cuisine', 'Specialties', 'Website']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for idx, restaurant in enumerate(restaurants, 1):
        row = idx + 1

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=restaurant.get('name', ''))
        ws.cell(row=row, column=3, value=restaurant.get('city', ''))
        ws.cell(row=row, column=4, value=restaurant.get('lga', ''))
        ws.cell(row=row, column=5, value=restaurant.get('location', ''))
        ws.cell(row=row, column=6, value=restaurant.get('rating', ''))
        ws.cell(row=row, column=7, value=restaurant.get('cuisine', ''))

        # Join specialties
        specialties = restaurant.get('specialties', [])
        ws.cell(row=row, column=8, value=', '.join(specialties) if specialties else '')

        ws.cell(row=row, column=9, value=restaurant.get('url', ''))

        # Center align number and rating
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(1, len(restaurants) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 50

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add autofilter
    ws.auto_filter.ref = f'A1:I{len(restaurants) + 1}'


def create_lga_sheets(wb, restaurants, state):
    """Create separate sheet for each LGA"""
    # Group by LGA
    lga_groups = {}
    for r in restaurants:
        lga = r.get('lga') or 'Unknown'
        if lga not in lga_groups:
            lga_groups[lga] = []
        lga_groups[lga].append(r)

    # Create sheet for each LGA - None-safe sorting
    sorted_lgas = sorted(lga_groups.keys(), key=lambda x: x if x else 'ZZZ')
    for lga in sorted_lgas:
        lga_restaurants = lga_groups[lga]

        # Sheet name (max 31 chars for Excel)
        sheet_name = lga[:28] if len(lga) > 28 else lga
        ws = wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = f"{lga} Restaurants ({len(lga_restaurants)})"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25

        # Headers
        headers = ['#', 'Restaurant Name', 'City', 'Rating', 'Cuisine', 'Specialties', 'Address', 'Website']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for idx, restaurant in enumerate(lga_restaurants, 1):
            row = idx + 2

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=restaurant.get('name', ''))
            ws.cell(row=row, column=3, value=restaurant.get('city', ''))
            ws.cell(row=row, column=4, value=restaurant.get('rating', ''))
            ws.cell(row=row, column=5, value=restaurant.get('cuisine', ''))

            specialties = restaurant.get('specialties', [])
            ws.cell(row=row, column=6, value=', '.join(specialties) if specialties else '')

            ws.cell(row=row, column=7, value=restaurant.get('location', ''))
            ws.cell(row=row, column=8, value=restaurant.get('url', ''))

            # Center align
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(2, len(lga_restaurants) + 3):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 50

        # Freeze headers
        ws.freeze_panes = 'A3'

        # Add filter
        ws.auto_filter.ref = f'A2:H{len(lga_restaurants) + 2}'
